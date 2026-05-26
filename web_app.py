"""
新闻采编统计工具 - Web 版
Flask 后端，完美复刻桌面版全部功能
"""
import logging
import os
import json
import sys
import uuid

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, send_file

# ══════════════════════════════════════════════════════
#  AppData 路径 & 配置
# ══════════════════════════════════════════════════════
def get_app_dir() -> str:
    if os.environ.get("RENDER") or os.environ.get("RAILWAY_ENVIRONMENT"):
        d = os.path.join("/tmp", "news_stats")
    elif sys.platform == "win32":
        d = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "新闻采编统计工具")
    else:
        d = os.path.join(os.path.expanduser("~"), ".news_stats")
    os.makedirs(d, exist_ok=True)
    return d

APP_DIR = get_app_dir()
CONFIG_FILE = os.path.join(APP_DIR, "config.json")

DEFAULT_CONFIG = {
    "data_sheet": "稿件收集",
    "index_sheet": "新闻指标档案",
    "output_sheet": "统计汇总",
    "special_persons": ["杨简", "韩友芳"],
    "rate_high": 80,
    "rate_mid": 50,
    "auto_open": False,
}

_log_file = os.path.join(APP_DIR, "app.log")
try:
    logging.basicConfig(
        level=logging.WARNING,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.FileHandler(_log_file, encoding="utf-8"), logging.StreamHandler()]
    )
except Exception:
    logging.basicConfig(level=logging.WARNING)


def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                d.setdefault(k, v)
            return d
        except (json.JSONDecodeError, IOError, OSError, PermissionError) as e:
            logging.error(f"配置文件读取失败: {e}")
    return dict(DEFAULT_CONFIG)


def save_config(cfg: dict) -> None:
    _tmp = CONFIG_FILE + ".tmp"
    try:
        with open(_tmp, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        os.replace(_tmp, CONFIG_FILE)
    except Exception as e:
        if os.path.exists(_tmp):
            os.remove(_tmp)
        raise


# ══════════════════════════════════════════════════════
#  Excel 工具
# ══════════════════════════════════════════════════════
def _rate_formula(col_e: str, col_f: str) -> str:
    return f'=IF({col_f}="","",IF({col_f}=0,"",{col_e}/{col_f}))'


def _safe_save_workbook(wb, filepath: str) -> None:
    _tmp = filepath + ".tmp"
    try:
        wb.save(_tmp)
        os.replace(_tmp, filepath)
    except Exception as exc:
        if os.path.exists(_tmp):
            os.remove(_tmp)
        raise


# ══════════════════════════════════════════════════════
#  核心统计逻辑
# ══════════════════════════════════════════════════════
def generate_summary(filepath: str, cfg: dict):
    df1 = pd.read_excel(filepath, sheet_name=cfg["data_sheet"])
    df3 = pd.read_excel(filepath, sheet_name=cfg["index_sheet"])
    SPECIAL = set(cfg["special_persons"])

    # 数据预清洗
    clean_cols = ["人员", "栏目", "成稿/线索", "新闻相关部门"]
    df1 = df1.copy()
    for col in clean_cols:
        if col in df1.columns:
            df1[col] = df1[col].astype(str).str.strip().replace(['nan', 'None', ''], [None, None, None])

    # 四维匹配差异化
    df1["_temp_dept"] = df1.apply(
        lambda x: x["新闻相关部门"] if x["人员"] in SPECIAL else None, axis=1
    )

    grouped = df1.groupby(
        ["人员", "栏目", "成稿/线索", "_temp_dept"], dropna=False
    ).agg(汇总数=("新闻内容", "count")).reset_index()
    grouped = grouped.rename(columns={"_temp_dept": "新闻相关部门"})

    def lookup(row):
        p, c, t, d = row["人员"], row["栏目"], row["成稿/线索"], row["新闻相关部门"]
        if p in SPECIAL:
            m = df3[(df3["媒体通联"] == p) & (df3["栏目"] == c) &
                    (df3["成稿/线索"] == t) & (df3["新闻相关部门"] == d)]
        else:
            m = df3[(df3["媒体通联"] == p) & (df3["栏目"] == c) & (df3["成稿/线索"] == t)]
        return m["指标"].values[0] if len(m) > 0 else None

    grouped["指标"] = grouped.apply(lookup, axis=1)
    grouped = grouped.sort_values(["人员", "栏目"]).reset_index(drop=True)

    # 写 Excel 统计汇总
    wb = load_workbook(filepath)
    if cfg["output_sheet"] in wb.sheetnames:
        del wb[cfg["output_sheet"]]
    ws = wb.create_sheet(cfg["output_sheet"])

    HDR_FILL = PatternFill("solid", start_color="2C2C2A", end_color="2C2C2A")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    ALT_FILL = PatternFill("solid", start_color="F7F7F5", end_color="F7F7F5")
    BD = Border(
        left=Side(style="thin", color="E8E8E4"), right=Side(style="thin", color="E8E8E4"),
        top=Side(style="thin", color="E8E8E4"), bottom=Side(style="thin", color="E8E8E4"),
    )
    CA = Alignment(horizontal="center", vertical="center")

    headers    = ["人员","栏目","成稿/线索","新闻相关部门","汇总数","指标","达成率"]
    col_widths = [12, 18, 12, 16, 10, 10, 12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CA; cell.border = BD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    for ri, row in grouped.iterrows():
        er = ri + 2
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        vals = [row["人员"], row["栏目"], row["成稿/线索"],
                row["新闻相关部门"] if pd.notna(row["新闻相关部门"]) else None,
                int(row["汇总数"]),
                row["指标"] if pd.notna(row["指标"]) else None]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=er, column=ci, value=val)
            c.fill = fill; c.alignment = CA; c.border = BD
            c.font = Font(name="Arial", size=10)
        c = ws.cell(row=er, column=7, value=_rate_formula(f"E{er}", f"F{er}"))
        c.fill = fill; c.alignment = CA; c.border = BD
        c.font = Font(name="Arial", size=10); c.number_format = "0.0%"

    tr = len(grouped) + 2
    for ci in range(1, 8):
        c = ws.cell(row=tr, column=ci)
        c.fill = PatternFill("solid", start_color="ECECEA", end_color="ECECEA")
        c.border = BD; c.alignment = CA
        c.font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=tr, column=1, value="合计")
    ws.cell(row=tr, column=5, value=f"=SUM(E2:E{tr-1})")
    ws.freeze_panes = "A2"

    # 达成进度
    idx_by_person = df3.groupby("媒体通联")["指标"].sum().reset_index()
    idx_by_person.columns = ["媒体通联", "指标总数"]
    submit_by_person = df1.groupby("人员").size().reset_index(name="提报总数")
    submit_by_person.columns = ["媒体通联", "提报总数"]
    publish_by_person = df1[df1["已刊发"].fillna(False)].groupby("人员").size().reset_index(name="刊发总数")
    publish_by_person.columns = ["媒体通联", "刊发总数"]

    progress_df = idx_by_person.merge(submit_by_person, on="媒体通联", how="left")
    progress_df = progress_df.merge(publish_by_person, on="媒体通联", how="left")
    progress_df["提报总数"] = progress_df["提报总数"].fillna(0).astype(int)
    progress_df["刊发总数"] = progress_df["刊发总数"].fillna(0).astype(int)
    progress_df["提报达成率"] = progress_df["提报总数"] / progress_df["指标总数"]
    progress_df = progress_df.sort_values("提报达成率", ascending=False).reset_index(drop=True)

    PROG_SHEET = "达成进度"
    if PROG_SHEET in wb.sheetnames:
        del wb[PROG_SHEET]
    wp = wb.create_sheet(PROG_SHEET)
    p_headers = ["媒体通联", "指标总数", "提报总数", "提报达成率", "刊发总数", "刊发率"]
    p_col_widths = [14, 12, 12, 14, 12, 12]
    for ci, (h, w) in enumerate(zip(p_headers, p_col_widths), 1):
        cell = wp.cell(row=1, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CA; cell.border = BD
        wp.column_dimensions[get_column_letter(ci)].width = w
    wp.row_dimensions[1].height = 22

    for ri, row in progress_df.iterrows():
        er = ri + 2
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        for ci, val in enumerate([row["媒体通联"], int(row["指标总数"]), int(row["提报总数"])], 1):
            c = wp.cell(row=er, column=ci, value=val)
            c.fill = fill; c.alignment = CA; c.border = BD
            c.font = Font(name="Arial", size=10)
        c = wp.cell(row=er, column=4, value=_rate_formula(f"C{er}", f"B{er}"))
        c.fill = fill; c.alignment = CA; c.border = BD
        c.font = Font(name="Arial", size=10); c.number_format = "0.0%"
        c = wp.cell(row=er, column=5, value=int(row["刊发总数"]))
        c.fill = fill; c.alignment = CA; c.border = BD
        c.font = Font(name="Arial", size=10)
        c = wp.cell(row=er, column=6, value=_rate_formula(f"E{er}", f"C{er}"))
        c.fill = fill; c.alignment = CA; c.border = BD
        c.font = Font(name="Arial", size=10); c.number_format = "0.0%"
    wp.freeze_panes = "A2"

    # 编辑发主编统计
    EDITOR_SHEET = "编辑发主编统计"
    if EDITOR_SHEET in wb.sheetnames:
        del wb[EDITOR_SHEET]
    we = wb.create_sheet(EDITOR_SHEET)
    e_headers = ["编辑", "发主编总数"]
    e_col_widths = [16, 14]
    for ci, (h, w) in enumerate(zip(e_headers, e_col_widths), 1):
        cell = we.cell(row=1, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CA; cell.border = BD
        we.column_dimensions[get_column_letter(ci)].width = w
    we.row_dimensions[1].height = 22

    df_edit = pd.read_excel(filepath, sheet_name="编辑改稿")
    sent = df_edit[df_edit["发主编"] == "发主编"].copy()
    by_editor = sent.groupby("编辑").size().reset_index(name="发主编总数")
    by_editor = by_editor.sort_values("发主编总数", ascending=False).reset_index(drop=True)

    for ri, row in by_editor.iterrows():
        er = ri + 2
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        for ci, val in enumerate([row["编辑"], int(row["发主编总数"])], 1):
            c = we.cell(row=er, column=ci, value=val)
            c.fill = fill; c.alignment = CA; c.border = BD
            c.font = Font(name="Arial", size=10)

    tr_e = len(by_editor) + 2
    for ci in range(1, 3):
        c = we.cell(row=tr_e, column=ci)
        c.fill = PatternFill("solid", start_color="ECECEA", end_color="ECECEA")
        c.border = BD; c.alignment = CA
        c.font = Font(name="Arial", bold=True, size=10)
    we.cell(row=tr_e, column=1, value="合计")
    we.cell(row=tr_e, column=2, value=f"=SUM(B2:B{tr_e-1})")
    we.freeze_panes = "A2"

    _safe_save_workbook(wb, filepath)

    awards = _calc_awards(df1, grouped, df_edit, progress_df)
    return grouped, progress_df, awards, by_editor


def _top3_distinct(series_sorted: pd.DataFrame, name_col: str, val_col: str,
                   extra_cols: list | None = None) -> list[dict]:
    result = []
    seen_vals = []
    for _, row in series_sorted.iterrows():
        v = round(float(row[val_col]), 6)
        if v not in seen_vals:
            seen_vals.append(v)
        rank = seen_vals.index(v) + 1
        if rank > 3:
            break
        entry = {"name": row[name_col], "val": float(row[val_col]), "rank": rank}
        if extra_cols:
            for col in extra_cols:
                entry[col] = row[col]
        result.append(entry)
        if len(seen_vals) >= 3:
            break
    return result


def _calc_awards(df1: pd.DataFrame, grouped: pd.DataFrame,
                 df_edit: pd.DataFrame, progress_df: pd.DataFrame | None) -> dict:
    if progress_df is not None and not progress_df.empty:
        top_rates = _top3_distinct(progress_df, "媒体通联", "提报达成率")
    else:
        top_rates = []

    person_submit = df1.groupby("人员").size().reset_index(name="总提报")
    person_pub = df1[df1["已刊发"].fillna(False)].groupby("人员").size().reset_index(name="已刊发数")
    hit = person_submit.merge(person_pub, on="人员", how="left")
    hit["已刊发数"] = hit["已刊发数"].fillna(0).astype(int)
    hit["刊发率"] = hit["已刊发数"] / hit["总提报"]
    hit = hit[hit["总提报"] >= 2].sort_values("刊发率", ascending=False).reset_index(drop=True)

    if not hit.empty:
        top_hits = _top3_distinct(hit, "人员", "刊发率", extra_cols=["已刊发数", "总提报"])
    else:
        top_hits = []

    df_sent = df_edit[df_edit["发主编"] == "发主编"].copy()
    if not df_sent.empty:
        editor_count = df_sent.groupby("编辑").size().reset_index(name="编稿数")
        editor_count = editor_count.sort_values("编稿数", ascending=False).reset_index(drop=True)
        top_editors = []
        for _, r in editor_count.head(2).iterrows():
            top_editors.append({"name": r["编辑"], "count": int(r["编稿数"])})
        while len(top_editors) < 2:
            top_editors.append({"name": "—", "count": 0})
    else:
        top_editors = [{"name": "—", "count": 0}, {"name": "—", "count": 0}]

    return {"top_rates": top_rates, "top_hits": top_hits, "top_editors": top_editors}


# ══════════════════════════════════════════════════════
#  JSON 序列化辅助
# ══════════════════════════════════════════════════════
def _json_safe(obj):
    if isinstance(obj, pd.DataFrame):
        return obj.where(pd.notna(obj), None).to_dict(orient="records")
    if isinstance(obj, list):
        return [_json_safe(item) for item in obj]
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if hasattr(obj, "item"):
        return obj.item()
    if pd.isna(obj):
        return None
    return obj


# ══════════════════════════════════════════════════════
#  Flask 应用
# ══════════════════════════════════════════════════════
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "news-stats-secret-key-2024")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

UPLOAD_DIR = os.path.join(APP_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

_results_cache: dict[str, dict] = {}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config", methods=["GET"])
def api_get_config():
    return jsonify(load_config())


@app.route("/api/config", methods=["POST"])
def api_save_config():
    try:
        data = request.get_json()
        cfg = load_config()
        cfg.update({
            "data_sheet": data.get("data_sheet", cfg["data_sheet"]).strip(),
            "index_sheet": data.get("index_sheet", cfg["index_sheet"]).strip(),
            "output_sheet": data.get("output_sheet", cfg["output_sheet"]).strip(),
            "special_persons": [s.strip() for s in data.get("special_persons", "").split("、") if s.strip()],
            "rate_high": int(data.get("rate_high", cfg["rate_high"])),
            "rate_mid": int(data.get("rate_mid", cfg["rate_mid"])),
            "auto_open": bool(data.get("auto_open", cfg.get("auto_open", False))),
        })
        save_config(cfg)
        return jsonify({"status": "ok", "config": cfg})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"status": "error", "message": "未选择文件"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"status": "error", "message": "未选择文件"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "message": "请上传 .xlsx 或 .xls 格式的 Excel 文件"}), 400

    session_id = str(uuid.uuid4())
    safe_name = f"{session_id}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, safe_name)
    file.save(filepath)

    cfg = load_config()
    try:
        grouped_df, progress_df, awards, editor_df = generate_summary(filepath, cfg)

        grouped_records = grouped_df.copy()
        grouped_records["达成率值"] = grouped_records.apply(
            lambda r: round(float(r["汇总数"]) / float(r["指标"]), 4)
            if pd.notna(r["指标"]) and r["指标"] > 0 else None, axis=1
        )

        progress_records = progress_df.copy()
        progress_records["刊发率值"] = progress_records.apply(
            lambda r: round(float(r["刊发总数"]) / float(r["提报总数"]), 4)
            if r["提报总数"] > 0 else 0.0, axis=1
        )

        result = {
            "status": "ok",
            "session_id": session_id,
            "filepath": filepath,
            "filename": file.filename,
            "summary": _json_safe(grouped_records),
            "summary_count": len(grouped_df),
            "progress": _json_safe(progress_records),
            "progress_count": len(progress_df),
            "awards": _json_safe(awards),
            "editor": _json_safe(editor_df),
            "editor_count": len(editor_df),
            "editor_total": int(editor_df["发主编总数"].sum()) if len(editor_df) > 0 else 0,
        }
        _results_cache[session_id] = result
        return jsonify(result)
    except Exception as e:
        logging.error(f"统计失败: {e}", exc_info=True)
        try:
            os.remove(filepath)
        except Exception:
            pass
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/download/<session_id>")
def api_download(session_id):
    result = _results_cache.get(session_id)
    if not result:
        return jsonify({"status": "error", "message": "会话已过期，请重新上传文件"}), 404
    filepath = result.get("filepath")
    filename = result.get("filename", "统计结果.xlsx")
    if not filepath or not os.path.exists(filepath):
        return jsonify({"status": "error", "message": "文件已不存在，请重新上传"}), 404
    return send_file(
        filepath, as_attachment=True,
        download_name=f"统计结果_{filename}",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("=" * 50)
    print("  新闻采编统计工具 - Web 版")
    print(f"  访问: http://127.0.0.1:{port}")
    print("=" * 50)
    app.run(debug=True, host="0.0.0.0", port=port)
